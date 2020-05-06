import pandas as pd
import datetime
import argparse
from openpyxl import load_workbook

parser = argparse.ArgumentParser(description='Gera xls de peças')
parser.add_argument('data', type=str, help='a data da planilha a ser usada')
args = parser.parse_args()



print("Iniciando o programa")
data_atual = datetime.date.today()
if data_atual.month == 2 and data_atual.day == 29:
    data_atual = data_atual.replace(day=28)
data_limite = data_atual.replace(year=2018)
print("Data de vencimento das Tales: {}".format(data_limite))

data_planilha = args.data
print("Data da planilha: {}".format(data_planilha))
base_instalada_path = "{}/Base Instalada.xlsx".format(data_planilha)
consumiveis_path = "{}/Consumíveis.xlsx".format(data_planilha)

base_instalada = pd.read_excel(base_instalada_path, sheet_name='IB MR NOR CDG')
print("Tamanho da base instalada: {}".format(base_instalada.shape[0]))
consumiveis = pd.read_excel(consumiveis_path, sheet_name='Planilha1')

base_instalada_tales = base_instalada[base_instalada['TALES'] == 'Y']
print("Dessas {} tem tales".format(base_instalada_tales.shape[0]))
consumiveis_tales = consumiveis[(consumiveis['Material'] == 10018247) | (consumiveis['Material'] == 7391886)]

consumiveis_tales_novas = consumiveis_tales[consumiveis_tales['Dt.Consumo'].dt.date > data_limite]
print("Das Tales substituidas {} estão no prazo".format(consumiveis_tales_novas.shape[0]))

consumiveis_tales_novas_destination = consumiveis_tales_novas[(consumiveis_tales_novas['Destinação Mat.'] == 'CB - CONSUMED BILLABLE') |
                                                  (consumiveis_tales_novas['Destinação Mat.'] == 'C - CONSUMED')]
print("Das Tales novas {} foram consumidas".format(consumiveis_tales_novas_destination.shape[0]))

#teste = base_instalada_tales[consumiveis_tales_novas_destination['Nº série'] in base_instalada_tales['Nº de série']]
base_tales_vencida = base_instalada_tales[~base_instalada_tales['Nº de série'].isin(consumiveis_tales_novas_destination['Nº série'])]
print("Número de Tales Vencidas: {}".format(base_tales_vencida.shape[0]))

numero_por_estado = base_tales_vencida.groupby(['Rg'])['TALES'].count()
numero_por_estado['Total Vencida'] = base_tales_vencida.shape[0]
numero_por_estado['Base Tales'] = base_instalada_tales.shape[0]
numero_por_estado['Base Total'] = base_instalada.shape[0]

numero_por_estado_data = numero_por_estado.rename(data_planilha)
#teste.to_excel('{}/resultados.xlsx'.format(data_planilha))


arquivo = load_workbook('resultados1.xlsx')
planilha = arquivo.active
print(planilha)
for cel in planilha.iter_rows(max_col=1):
    # o retorno eh um tuple, com o primeiro valor sendo a celula
    cel = cel[0]
    print(cel.value)




"""
for ws in planilha.worksheets:
    for index, row in enumerate(ws.rows, start=1):
        ws.cell(row=index, column=2).value='oi'
# tem que ter um save aqui

"""